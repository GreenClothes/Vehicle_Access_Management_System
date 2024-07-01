import openpyxl
import numpy as np
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook

# 오늘 날짜로 된 파일이 있는지 확인
def FILE_CHECK(date = datetime.today().strftime("%Y%m%d")):
    file_path = './DATA/{0}.xlsx'.format(date)

    if not os.path.isfile(file_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        subject = ['TIME', 'LICENSE PLATE NUMBER', 'TYPE', 'APT NUM']
        sheet.append(subject)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 30
        wb.save(file_path)

    return file_path

# 등록 번호판 확인
def PLATE_CHAR_CHECK(char):
    file_path = 'DATA/plate_num.xlsx'
    plate_file = load_workbook(file_path)
    sheet = plate_file.active

    for i in range(1, len(list(sheet.rows))+1):
        if char == sheet.cell(i, 2).value:
            return '등록', int(sheet.cell(i, 1).value)

    return '방문', ''

# 번호판 인식 데이터 기록
def DATA_INPUT(char, DATA_TYPE):
    file_path = FILE_CHECK()
    file = load_workbook(file_path)
    sheet = file.active

    if DATA_TYPE == 'plate num':
        TYPE, APT_NUM = PLATE_CHAR_CHECK(char)
        in_DATA = ['{0}'.format(datetime.today().strftime("%H%M%S")), char, TYPE, APT_NUM]
        sheet.append(in_DATA)
        file.save(file_path)
        if APT_NUM == '':
            print('no')
            # 동/호수 정보 ATMEGA에서 받아오기 요청 및 받아오기
            return 0
        else:
            return 1
    elif DATA_TYPE == "can't detection":
        in_DATA = ['{0}'.format(datetime.today().strftime("%H%M%S")), 'No detection', 'Unknown', '']
        sheet.append(in_DATA)
        file.save(file_path)
    elif DATA_TYPE == 'apt num':
        sheet.cell(len(list(sheet.rows)), 4, char)
        file.save(file_path)

    return 0


# 번호판 인식 데이터 출력
def DATA_OUTPUT(date):
    file_path = FILE_CHECK(date)
    file = load_workbook(file_path)
    sheet = file.active
    date_DATA = []

    for row in range(2, len(list(sheet.rows))+1):
        date_DATA_cache = []
        for col in range(1, len(list(sheet.columns))+1):
            date_DATA_cache.append(sheet.cell(row=row, column=col).value)
        date_DATA.append(tuple(date_DATA_cache))

    return date_DATA
